from flask import Flask, render_template, request, jsonify, send_file, session
from flask_cors import CORS
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import time
import secrets
from werkzeug.utils import secure_filename
from datetime import datetime
import json

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)
CORS(app)

# Configuration
UPLOAD_FOLDER = 'uploads'
TEMPLATE_FOLDER = 'template'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Create necessary folders
for folder in [UPLOAD_FOLDER, TEMPLATE_FOLDER, OUTPUT_FOLDER]:
    os.makedirs(folder, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['TEMPLATE_FOLDER'] = TEMPLATE_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_template_file():
    """Get the current template file path"""
    template_files = [f for f in os.listdir(TEMPLATE_FOLDER) if allowed_file(f)]
    if template_files:
        return os.path.join(TEMPLATE_FOLDER, template_files[0])
    return None

def merge_excel_files(file1_path, file2_path):
    """Merge two Excel files vertically"""
    try:
        df1 = pd.read_excel(file1_path)
        df2 = pd.read_excel(file2_path)
        merged_df = pd.concat([df1, df2], ignore_index=True)
        return merged_df, None
    except Exception as e:
        return None, str(e)

def update_ga_raw_sheet(merged_data, template_path):
    """Update the GA Raw sheet in the template file"""
    try:
        wb = openpyxl.load_workbook(template_path)
        
        # Remove existing GA Raw sheet if it exists
        if 'GA Raw' in wb.sheetnames:
            del wb['GA Raw']
        
        # Create new GA Raw sheet at the beginning
        ws_raw = wb.create_sheet('GA Raw', 0)
        
        # Write data to GA Raw
        for r_idx, row in enumerate(dataframe_to_rows(merged_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_raw.cell(row=r_idx, column=c_idx, value=value)
        
        # Save the workbook
        wb.save(template_path)
        wb.close()
        
        # Wait for formulas to recalculate
        time.sleep(2)
        
        return True, None
    except Exception as e:
        return False, str(e)

def extract_cells_from_live_ga(template_path, cell_range):
    """Extract specific cells from Live GA sheet"""
    try:
        # Reopen workbook with calculated values
        wb = openpyxl.load_workbook(template_path, data_only=True)
        
        if 'Live GA' not in wb.sheetnames:
            return None, "'Live GA' sheet not found in template"
        
        ws_live = wb['Live GA']
        
        # Extract the specified range
        cell_data = []
        for row in ws_live[cell_range]:
            row_data = []
            for cell in row:
                value = cell.value if cell.value is not None else ""
                row_data.append(str(value))
            cell_data.append(row_data)
        
        wb.close()
        
        return cell_data, None
    except Exception as e:
        return None, str(e)

def format_cells_as_text(cell_data):
    """Format cell data as text for copying"""
    if not cell_data:
        return ""
    
    formatted_text = ""
    for row in cell_data:
        formatted_text += "\t".join(row) + "\n"
    
    return formatted_text.strip()

@app.route('/')
def index():
    """Render the main page"""
    template_exists = get_template_file() is not None
    template_info = None
    
    if template_exists:
        template_path = get_template_file()
        template_info = {
            'filename': os.path.basename(template_path),
            'uploaded': datetime.fromtimestamp(os.path.getmtime(template_path)).strftime('%Y-%m-%d %H:%M:%S')
        }
    
    return render_template('index.html', 
                         template_exists=template_exists,
                         template_info=template_info)

@app.route('/upload-template', methods=['POST'])
def upload_template():
    """Upload or replace the GA template file"""
    if 'template' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['template']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Invalid file type. Use .xlsx or .xls'}), 400
    
    try:
        # Clear existing template files
        for f in os.listdir(TEMPLATE_FOLDER):
            os.remove(os.path.join(TEMPLATE_FOLDER, f))
        
        # Save new template
        filename = secure_filename(file.filename)
        filepath = os.path.join(TEMPLATE_FOLDER, filename)
        file.save(filepath)
        
        # Verify it has the required sheets
        wb = openpyxl.load_workbook(filepath)
        sheets = wb.sheetnames
        wb.close()
        
        if 'Live GA' not in sheets:
            os.remove(filepath)
            return jsonify({
                'success': False, 
                'error': 'Template must contain a "Live GA" sheet'
            }), 400
        
        return jsonify({
            'success': True,
            'filename': filename,
            'sheets': sheets
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/delete-template', methods=['POST'])
def delete_template():
    """Delete the current template file"""
    try:
        for f in os.listdir(TEMPLATE_FOLDER):
            os.remove(os.path.join(TEMPLATE_FOLDER, f))
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/process', methods=['POST'])
def process_files():
    """Process the uploaded files and merge them"""
    if 'file1' not in request.files or 'file2' not in request.files:
        return jsonify({'success': False, 'error': 'Both files are required'}), 400
    
    file1 = request.files['file1']
    file2 = request.files['file2']
    cell_range = request.form.get('cellRange', 'A1:D10')
    
    if not file1.filename or not file2.filename:
        return jsonify({'success': False, 'error': 'Both files must be selected'}), 400
    
    if not allowed_file(file1.filename) or not allowed_file(file2.filename):
        return jsonify({'success': False, 'error': 'Invalid file type'}), 400
    
    template_path = get_template_file()
    if not template_path:
        return jsonify({'success': False, 'error': 'No template file uploaded. Please upload a template first.'}), 400
    
    try:
        # Save uploaded files temporarily
        file1_path = os.path.join(UPLOAD_FOLDER, secure_filename(file1.filename))
        file2_path = os.path.join(UPLOAD_FOLDER, secure_filename(file2.filename))
        
        file1.save(file1_path)
        file2.save(file2_path)
        
        # Step 1: Merge files
        merged_data, error = merge_excel_files(file1_path, file2_path)
        if error:
            return jsonify({'success': False, 'error': f'Merge failed: {error}'}), 500
        
        # Step 2: Update GA Raw sheet
        success, error = update_ga_raw_sheet(merged_data, template_path)
        if not success:
            return jsonify({'success': False, 'error': f'Update failed: {error}'}), 500
        
        # Step 3: Extract cells from Live GA
        cell_data, error = extract_cells_from_live_ga(template_path, cell_range)
        if error:
            return jsonify({'success': False, 'error': f'Extraction failed: {error}'}), 500
        
        # Step 4: Format as text
        formatted_text = format_cells_as_text(cell_data)
        
        # Clean up uploaded files
        os.remove(file1_path)
        os.remove(file2_path)
        
        return jsonify({
            'success': True,
            'data': cell_data,
            'formatted_text': formatted_text,
            'row_count': len(merged_data)
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get-template-info', methods=['GET'])
def get_template_info():
    """Get information about the current template"""
    template_path = get_template_file()
    
    if not template_path:
        return jsonify({'exists': False})
    
    try:
        wb = openpyxl.load_workbook(template_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        
        return jsonify({
            'exists': True,
            'filename': os.path.basename(template_path),
            'sheets': sheets,
            'uploaded': datetime.fromtimestamp(os.path.getmtime(template_path)).strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        return jsonify({'exists': False, 'error': str(e)})

@app.route('/download-template', methods=['GET'])
def download_template():
    """Download the current template file"""
    template_path = get_template_file()
    
    if not template_path:
        return jsonify({'error': 'No template file available'}), 404
    
    return send_file(template_path, as_attachment=True)

if __name__ == '__main__':
    print("\n" + "="*60)
    print("🚀 GA Automation Web Server Starting...")
    print("="*60)
    print("\n📊 Access the application at: http://localhost:5000")
    print("\n💡 Features:")
    print("   • Upload GA template (stays until you replace it)")
    print("   • Drag & drop File 1 and File 2 to merge")
    print("   • Auto-extract cells from Live GA")
    print("   • One-click copy to clipboard")
    print("\n" + "="*60 + "\n")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
